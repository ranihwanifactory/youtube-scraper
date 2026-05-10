import streamlit as st
import time
import random
import re
import json
import io
import pandas as pd
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

# 페이지 설정
st.set_page_config(
    page_title="유튜브 콘텐츠 스크래퍼",
    page_icon="🎬",
    layout="wide"
)

st.title("🎬 유튜브 콘텐츠 스크래퍼")
st.markdown("키워드 또는 채널을 입력하면 유튜브 콘텐츠를 자동으로 수집하고 AI가 인기 키워드를 추천합니다.")

# ── session_state 초기화 ───────────────────────────────
if 'df' not in st.session_state:
    st.session_state.df = None
if 'search_keyword' not in st.session_state:
    st.session_state.search_keyword = ''
if 'total_collected' not in st.session_state:
    st.session_state.total_collected = 0
if 'filtered_count' not in st.session_state:
    st.session_state.filtered_count = 0
if 'channel_list' not in st.session_state:
    st.session_state.channel_list = []
if 'selected_channels' not in st.session_state:
    st.session_state.selected_channels = []
# ── 추천 채널 캐시 ────────────────────────────────────
if 'channel_recommendations' not in st.session_state:
    st.session_state.channel_recommendations = None
if 'channel_rec_keyword' not in st.session_state:
    st.session_state.channel_rec_keyword = ''
# ── 구독자 수 캐시 {channel_url: subscriber_count} ────
if 'subscriber_cache' not in st.session_state:
    st.session_state.subscriber_cache = {}

# ── 상수 ─────────────────────────────────────────────
UPLOAD_FILTER_MAP = {
    "전체":       "",
    "1일 이내":   "&sp=EgIIAQ%253D%253D",
    "1주일 이내": "&sp=EgIIAw%253D%253D",
    "1개월 이내": "&sp=EgIIBA%253D%253D",
    "6개월 이내": "",
    "1년 이내":   "",
}

UPLOAD_DAY_LIMIT = {
    "전체": 9999, "1일 이내": 1, "1주일 이내": 7,
    "1개월 이내": 30, "6개월 이내": 180, "1년 이내": 365,
}


# ── 채널 URL 정규화 유틸 ──────────────────────────────
def normalize_channel_url(raw: str) -> str:
    ch = raw.strip()
    if ch.startswith("https://") or ch.startswith("http://"):
        return ch.rstrip('/')
    elif ch.startswith("@"):
        return f"https://www.youtube.com/{ch}"
    else:
        return f"https://www.youtube.com/@{ch}"


# ── 사이드바 ──────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 검색 설정")

    search_mode = st.radio(
        "🎯 검색 모드",
        options=["키워드 검색", "채널 검색"],
        horizontal=True,
        help="키워드 검색: 유튜브 전체에서 키워드로 검색\n채널 검색: 등록한 채널 목록에서 수집"
    )

    st.divider()

    if search_mode == "키워드 검색":
        keyword = st.text_input("🔍 검색 키워드", placeholder="예: 성주참외, 골프 레슨...")
        channel_input = ""

    else:
        st.subheader("📺 채널 목록 관리")

        col_ch, col_add = st.columns([3, 1])
        with col_ch:
            new_ch_input = st.text_input(
                "채널 추가",
                placeholder="@채널명 또는 URL",
                label_visibility="collapsed",
                key="new_ch_input"
            )
        with col_add:
            st.markdown("<br>", unsafe_allow_html=True)
            add_btn = st.button("➕", use_container_width=True, key="add_ch_btn")

        if add_btn and new_ch_input.strip():
            normalized = normalize_channel_url(new_ch_input.strip())
            display_name = new_ch_input.strip()
            existing_urls = [c['url'] for c in st.session_state.channel_list]
            if normalized not in existing_urls:
                st.session_state.channel_list.append({
                    'name': display_name,
                    'url':  normalized
                })
                st.rerun()
            else:
                st.warning("이미 추가된 채널입니다.")

        if not st.session_state.channel_list:
            st.caption("아직 추가된 채널이 없습니다.")
        else:
            st.caption(f"총 {len(st.session_state.channel_list)}개 채널 등록됨")
            to_delete = None
            for i, ch in enumerate(st.session_state.channel_list):
                col_name, col_del = st.columns([4, 1])
                with col_name:
                    st.markdown(
                        f"<small>{'✅' if ch['url'] in st.session_state.selected_channels else '⬜'} "
                        f"{ch['name']}</small>",
                        unsafe_allow_html=True
                    )
                with col_del:
                    if st.button("🗑️", key=f"del_{i}", use_container_width=True):
                        to_delete = i
            if to_delete is not None:
                removed_url = st.session_state.channel_list[to_delete]['url']
                st.session_state.channel_list.pop(to_delete)
                chk_key = f"chk_{removed_url}"
                if chk_key in st.session_state:
                    del st.session_state[chk_key]
                if removed_url in st.session_state.selected_channels:
                    st.session_state.selected_channels.remove(removed_url)
                st.rerun()

        st.divider()

        st.subheader("✅ 수집할 채널 선택")
        if not st.session_state.channel_list:
            st.caption("위에서 채널을 먼저 추가하세요.")
            selected_channels = []
        else:
            col_all, col_none = st.columns(2)
            with col_all:
                if st.button("전체 선택", use_container_width=True, key="sel_all"):
                    for ch in st.session_state.channel_list:
                        st.session_state[f"chk_{ch['url']}"] = True
                    st.rerun()
            with col_none:
                if st.button("전체 해제", use_container_width=True, key="sel_none"):
                    for ch in st.session_state.channel_list:
                        st.session_state[f"chk_{ch['url']}"] = False
                    st.rerun()

            selected_channels = []
            for ch in st.session_state.channel_list:
                chk_key = f"chk_{ch['url']}"
                if chk_key not in st.session_state:
                    st.session_state[chk_key] = ch['url'] in st.session_state.selected_channels
                val = st.checkbox(ch['name'], key=chk_key)
                if val:
                    selected_channels.append(ch['url'])

            st.session_state.selected_channels = selected_channels

        st.divider()

        # ── 채널 검색 방식 선택 ──────────────────────────
        st.subheader("🔍 채널 내 검색 방식")
        channel_search_mode = st.radio(
            "검색 방식",
            options=["탭 수집 (전체 영상)", "키워드로 채널 내 검색"],
            help="탭 수집: 채널의 동영상/쇼츠 탭 전체 수집\n키워드 검색: 채널 내에서 특정 키워드로 검색",
            key="channel_search_mode"
        )

        if channel_search_mode == "탭 수집 (전체 영상)":
            st.subheader("📅 수집 탭")
            channel_tab = st.selectbox(
                "채널 페이지 탭",
                options=["동영상", "쇼츠"],
                help="동영상: 일반 영상 탭 / 쇼츠: 쇼츠 탭"
            )
            keyword = st.text_input(
                "🔍 제목 키워드 필터 (선택)",
                placeholder="비워두면 전체 영상 수집",
                help="입력 시 해당 단어가 제목에 포함된 영상만 표시합니다",
                key="ch_keyword"
            )
        else:
            channel_tab = "동영상"
            keyword = st.text_input(
                "🔍 채널 내 검색 키워드 (필수)",
                placeholder="예: 레시피, 브이로그...",
                help="채널 내에서 이 키워드로 검색합니다",
                key="ch_search_keyword"
            )

        channel_input = ""

    st.divider()

    st.subheader("🩳 콘텐츠 유형")
    content_type = st.radio(
        "수집할 영상 유형",
        options=["전체 (일반 + 쇼츠)", "일반 영상만", "쇼츠만"],
        index=0,
        help="쇼츠: 60초 이하 세로형 영상"
    )

    st.divider()

    if search_mode == "키워드 검색":
        st.subheader("📅 업로드 시간 필터")
        upload_filter = st.selectbox(
            "업로드 시간 범위",
            options=["전체", "1일 이내", "1주일 이내", "1개월 이내", "6개월 이내", "1년 이내"]
        )
    else:
        upload_filter = "전체"

    st.divider()

    st.subheader("👁️ 조회수 필터")
    use_view_filter = st.checkbox("최소 조회수 필터 사용")
    min_view = 0
    if use_view_filter:
        min_view = st.number_input(
            "최소 조회수 (이상만 수집)",
            min_value=0, value=10000, step=1000, format="%d"
        )
        st.caption(f"📌 {min_view:,}회 이상인 영상만 표시됩니다.")

    st.divider()
    run_btn = st.button("▶ 크롤링 시작", type="primary", use_container_width=True)
    st.caption("📌 크롤링 중 크롬 브라우저가 자동으로 열립니다.")
    st.caption("📌 결과는 CSV로 다운로드할 수 있습니다.")


# ── 유틸 함수 ─────────────────────────────────────────
def date_to_days(date_str: str) -> int:
    date_str = date_str.strip()
    patterns = [
        (r'(\d+)일 전',   lambda m: int(m.group(1))),
        (r'(\d+)주 전',   lambda m: int(m.group(1)) * 7),
        (r'(\d+)개월 전', lambda m: int(m.group(1)) * 30),
        (r'(\d+)년 전',   lambda m: int(m.group(1)) * 365),
        (r'(\d+)시간 전', lambda m: 0),
        (r'(\d+)분 전',   lambda m: 0),
    ]
    for pattern, calc in patterns:
        m = re.search(pattern, date_str)
        if m:
            return calc(m)
    return 9999


def parse_view_count(view_str: str) -> int:
    if not view_str:
        return 0
    s = view_str.strip()
    s = re.sub(r'[,\s•\n\r]', '', s)
    try:
        if '억' in s:
            eok = re.search(r'([\d.]+)억', s)
            man = re.search(r'억([\d.]+)만', s)
            result = 0
            if eok:
                result += float(eok.group(1)) * 100_000_000
            if man:
                result += float(man.group(1)) * 10_000
            return int(result) if result else 0
        if '만' in s:
            num = re.search(r'([\d.]+)만', s)
            return int(float(num.group(1)) * 10_000) if num else 0
        if '천' in s:
            num = re.search(r'([\d.]+)천', s)
            return int(float(num.group(1)) * 1_000) if num else 0
        pure = re.sub(r'[^\d]', '', s)
        return int(pure) if pure else 0
    except Exception:
        return 0


def is_shorts(link: str, title: str = "") -> bool:
    if '/shorts/' in link:
        return True
    if '#shorts' in title.lower() or '#short' in title.lower():
        return True
    return False


# ── 무한 스크롤 (lazy-load 대기 강화) ────────────────────
def scroll(driver, max_same_count: int = 3):
    same_count = 0
    last_h = driver.execute_script("return document.documentElement.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        time.sleep(random.uniform(1.5, 2.5))
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight - 200);")
        time.sleep(0.5)
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
        time.sleep(1.0)
        new_h = driver.execute_script("return document.documentElement.scrollHeight")
        if new_h == last_h:
            same_count += 1
            if same_count >= max_same_count:
                break
        else:
            same_count = 0
            last_h = new_h


def get_driver():
    import shutil, os
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--headless')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-setuid-sandbox')
    options.add_argument('--remote-debugging-port=9222')
    options.add_argument('--lang=ko-KR')
    options.add_argument('--accept-lang=ko-KR,ko;q=0.9')

    chromium_paths = [
        '/usr/bin/chromedriver',
        '/usr/lib/chromium-browser/chromedriver',
        shutil.which('chromedriver') or '',
    ]
    chromium_bin_paths = [
        '/usr/bin/chromium',
        '/usr/bin/chromium-browser',
        shutil.which('chromium') or '',
        shutil.which('chromium-browser') or '',
    ]

    driver_path = next((p for p in chromium_paths if p and os.path.exists(p)), None)
    binary_path = next((p for p in chromium_bin_paths if p and os.path.exists(p)), None)

    if driver_path:
        if binary_path:
            options.binary_location = binary_path
        service = Service(driver_path)
    else:
        service = Service(ChromeDriverManager().install())

    return webdriver.Chrome(service=service, options=options)


# ── ytInitialData JSON 추출 ──────────────────────────
def _extract_yt_initial_data(html: str) -> dict:
    patterns = [
        r'var ytInitialData\s*=\s*(\{.*?\});\s*</script>',
        r'window\["ytInitialData"\]\s*=\s*(\{.*?\});',
        r'ytInitialData\s*=\s*(\{.*?\});\s*(?:var|window|//)',
    ]
    for pat in patterns:
        m = re.search(pat, html, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(1))
            except Exception:
                continue
    return {}


def _safe_get(obj, *keys, default=''):
    for key in keys:
        if obj is None:
            return default
        if isinstance(obj, dict):
            obj = obj.get(key)
        elif isinstance(obj, list) and isinstance(key, int):
            obj = obj[key] if key < len(obj) else None
        else:
            return default
    return obj if obj is not None else default


def _parse_video_renderer(renderer: dict):
    if not isinstance(renderer, dict):
        return None

    video_id = (
        renderer.get('videoId') or
        _safe_get(renderer, 'navigationEndpoint', 'watchEndpoint', 'videoId') or
        _safe_get(renderer, 'navigationEndpoint', 'reelWatchEndpoint', 'videoId') or
        _safe_get(renderer, 'onClickCommand', 'reelWatchEndpoint', 'videoId') or
        ''
    )
    if not video_id:
        return None

    title = (
        _safe_get(renderer, 'title', 'runs', 0, 'text') or
        _safe_get(renderer, 'title', 'simpleText') or
        _safe_get(renderer, 'headline', 'simpleText') or
        _safe_get(renderer, 'headline', 'runs', 0, 'text') or
        _safe_get(renderer, 'accessibility', 'accessibilityData', 'label') or
        ''
    )

    view_raw = (
        _safe_get(renderer, 'viewCountText', 'simpleText') or
        _safe_get(renderer, 'viewCountText', 'runs', 0, 'text') or
        _safe_get(renderer, 'shortViewCountText', 'simpleText') or
        _safe_get(renderer, 'shortViewCountText', 'runs', 0, 'text') or
        _safe_get(renderer, 'videoInfo', 'runs', 0, 'text') or
        ''
    )

    date_raw = (
        _safe_get(renderer, 'publishedTimeText', 'simpleText') or
        _safe_get(renderer, 'publishedTimeText', 'runs', 0, 'text') or
        _safe_get(renderer, 'videoInfo', 'runs', 2, 'text') or
        _safe_get(renderer, 'videoInfo', 'runs', 4, 'text') or
        ''
    )

    # ── 채널명 추출 ──────────────────────────────────
    channel_name_raw = (
        _safe_get(renderer, 'ownerText', 'runs', 0, 'text') or
        _safe_get(renderer, 'longBylineText', 'runs', 0, 'text') or
        _safe_get(renderer, 'shortBylineText', 'runs', 0, 'text') or
        ''
    )
    channel_url_raw = (
        _safe_get(renderer, 'ownerText', 'runs', 0, 'navigationEndpoint', 'browseEndpoint', 'canonicalBaseUrl') or
        _safe_get(renderer, 'longBylineText', 'runs', 0, 'navigationEndpoint', 'browseEndpoint', 'canonicalBaseUrl') or
        _safe_get(renderer, 'shortBylineText', 'runs', 0, 'navigationEndpoint', 'browseEndpoint', 'canonicalBaseUrl') or
        ''
    )

    # ── 쇼츠 판별 ────────────────────────────────────
    is_reel = bool(
        _safe_get(renderer, 'navigationEndpoint', 'reelWatchEndpoint', 'videoId') or
        _safe_get(renderer, 'onClickCommand', 'reelWatchEndpoint', 'videoId')
    )
    length_text = (
        _safe_get(renderer, 'lengthText', 'simpleText') or
        _safe_get(renderer, 'lengthText', 'runs', 0, 'text') or
        ''
    )
    def _is_short_duration(lt: str) -> bool:
        if not lt:
            return False
        parts = lt.strip().split(':')
        try:
            if len(parts) == 2:
                return int(parts[0]) == 0
            if len(parts) == 1:
                return int(parts[0]) <= 60
        except Exception:
            pass
        return False

    link_url = (
        f'https://www.youtube.com/shorts/{video_id}'
        if is_reel or _is_short_duration(length_text)
        else f'https://www.youtube.com/watch?v={video_id}'
    )

    return {
        'title':        str(title).strip(),
        'link':         link_url,
        'view':         _extract_view(str(view_raw)),
        'upload_date':  str(date_raw).strip(),
        'channel_name': str(channel_name_raw).strip(),
        'channel_url':  f'https://www.youtube.com{channel_url_raw}' if channel_url_raw and not channel_url_raw.startswith('http') else str(channel_url_raw).strip(),
    }


def _parse_lockup_view_model(lvm: dict):
    if not isinstance(lvm, dict):
        return None

    video_id = lvm.get('contentId', '')
    if not video_id:
        try:
            sources = lvm['contentImage']['thumbnailViewModel']['image']['sources']
            url = sources[0].get('url', '') if sources else ''
            m = re.search(r'/vi/([a-zA-Z0-9_-]{11})/', url)
            if m:
                video_id = m.group(1)
        except Exception:
            pass
    if not video_id:
        return None

    title = (
        _safe_get(lvm, 'metadata', 'lockupMetadataViewModel', 'title', 'content') or
        _safe_get(lvm, 'title', 'content') or
        _safe_get(lvm, 'accessibilityText') or
        ''
    )

    view_raw = ''
    date_raw = ''
    try:
        lmvm = lvm.get('metadata', {}).get('lockupMetadataViewModel', {})
        cmvm = lmvm.get('metadata', {}).get('contentMetadataViewModel', {})
        rows_data = cmvm.get('metadataRows', [])
        if not rows_data:
            rows_data = lmvm.get('metadataRows', [])

        all_texts = []
        for row in rows_data:
            for part in row.get('metadataParts', []):
                txt = (
                    _safe_get(part, 'text', 'content') or
                    _safe_get(part, 'text', 'simpleText') or
                    _safe_get(part, 'text', 'runs', 0, 'text') or
                    ''
                )
                if txt:
                    all_texts.append(txt.strip())

        for txt in all_texts:
            if not view_raw and (
                '조회수' in txt or
                re.search(r'[\d.]+\s*[만천억]', txt) or
                re.search(r'\d{1,3}(?:,\d{3})+', txt) or
                re.search(r'^\d{4,}$', txt)
            ):
                view_raw = txt
            elif not date_raw and re.search(
                r'\d+\s*(?:분|시간|일|주|개월|년)\s*전', txt
            ):
                date_raw = txt

    except Exception:
        pass

    if not view_raw or not date_raw:
        acc = lvm.get('accessibilityText', '')
        if acc:
            if not view_raw:
                m_v = re.search(r'조회수\s*([\d.,]+\s*(?:[만천억])?)', acc)
                if m_v:
                    view_raw = m_v.group(0)
            if not date_raw:
                m_d = re.search(r'\d+\s*(?:분|시간|일|주|개월|년)\s*전', acc)
                if m_d:
                    date_raw = m_d.group(0)

    return {
        'title':        str(title).strip(),
        'link':         f'https://www.youtube.com/watch?v={video_id}',
        'view':         _extract_view(str(view_raw)),
        'upload_date':  str(date_raw).strip(),
        'channel_name': '',
        'channel_url':  '',
    }


def _walk_renderers(obj, key='videoRenderer', results=None):
    if results is None:
        results = []
    if isinstance(obj, dict):
        if key in obj:
            results.append(obj[key])
        for v in obj.values():
            _walk_renderers(v, key, results)
    elif isinstance(obj, list):
        for item in obj:
            _walk_renderers(item, key, results)
    return results


def _parse_channel_search_data(data: dict) -> list:
    rows = []
    item_sections = _walk_renderers(data, 'itemSectionRenderer')
    for section in item_sections:
        if not isinstance(section, dict):
            continue
        for item in section.get('contents', []):
            if not isinstance(item, dict):
                continue
            vr = item.get('videoRenderer')
            if vr:
                parsed = _parse_video_renderer(vr)
                if parsed:
                    rows.append(parsed)
    seen = set()
    unique = []
    for r in rows:
        vid = r['link'].split('/')[-1].split('?')[0].replace('watch?v=', '')
        if '=' in vid:
            vid = vid.split('=')[-1]
        if vid and vid not in seen:
            seen.add(vid)
            unique.append(r)
    return unique


def _parse_all_renderers_from_data(data: dict) -> list:
    rows = []

    rich_items = _walk_renderers(data, 'richItemRenderer')
    for ri in rich_items:
        inner_obj = ri.get('content') if isinstance(ri, dict) and 'content' in ri else ri
        if not isinstance(inner_obj, dict):
            continue

        handled = False
        for inner_key in ('videoRenderer', 'gridVideoRenderer',
                          'reelItemRenderer', 'shortsLockupViewModel'):
            inner_list = _walk_renderers(inner_obj, inner_key)
            if inner_list:
                handled = True
            for vr in inner_list:
                parsed = _parse_video_renderer(vr)
                if parsed:
                    rows.append(parsed)

        lvm = inner_obj.get('lockupViewModel')
        if lvm and not handled:
            parsed = _parse_lockup_view_model(lvm)
            if parsed:
                rows.append(parsed)

    for vr in _walk_renderers(data, 'gridVideoRenderer'):
        parsed = _parse_video_renderer(vr)
        if parsed:
            rows.append(parsed)

    for vr in _walk_renderers(data, 'videoRenderer'):
        parsed = _parse_video_renderer(vr)
        if parsed:
            rows.append(parsed)

    for vr in _walk_renderers(data, 'reelItemRenderer'):
        parsed = _parse_video_renderer(vr)
        if parsed:
            rows.append(parsed)

    for vr in _walk_renderers(data, 'shortsLockupViewModel'):
        vid = (
            _safe_get(vr, 'onTap', 'innertubeCommand', 'reelWatchEndpoint', 'videoId') or
            _safe_get(vr, 'entityId') or ''
        )
        if vid.startswith('shorts-shelf-item-'):
            vid = vid.replace('shorts-shelf-item-', '')
        title = (
            _safe_get(vr, 'overlayMetadata', 'primaryText', 'content') or
            _safe_get(vr, 'accessibilityText') or ''
        )
        view_raw = _safe_get(vr, 'overlayMetadata', 'secondaryText', 'content') or ''
        if vid:
            rows.append({
                'title':        str(title).strip(),
                'link':         f'https://www.youtube.com/watch?v={vid}',
                'view':         _extract_view(str(view_raw)),
                'upload_date':  '',
                'channel_name': '',
                'channel_url':  '',
            })

    seen = set()
    unique = []
    for r in rows:
        vid = r['link'].split('v=')[-1].split('&')[0]
        if vid and vid not in seen:
            seen.add(vid)
            unique.append(r)

    return unique


def _save_debug(key: str, url: str, rows: int, html: str, method: str):
    if 'scrape_debug' not in st.session_state:
        st.session_state.scrape_debug = {}
    st.session_state.scrape_debug[key] = {
        'url': url,
        'rows': rows,
        'has_ytInitialData': 'ytInitialData' in html,
        'html_len': len(html),
        'parse_method': method,
    }


# ── 채널 구독자 수 수집 ───────────────────────────────
def parse_subscriber_count(sub_str: str) -> int:
    """구독자 수 문자열을 정수로 변환합니다."""
    if not sub_str:
        return 0
    s = sub_str.strip()
    s = re.sub(r'[구독자\s,]', '', s)
    try:
        if '억' in s:
            m = re.search(r'([\d.]+)억', s)
            return int(float(m.group(1)) * 100_000_000) if m else 0
        if '만' in s:
            m = re.search(r'([\d.]+)만', s)
            return int(float(m.group(1)) * 10_000) if m else 0
        if '천' in s:
            m = re.search(r'([\d.]+)천', s)
            return int(float(m.group(1)) * 1_000) if m else 0
        # 영어권: 1.2M, 500K
        if 'M' in s.upper():
            m = re.search(r'([\d.]+)[Mm]', s)
            return int(float(m.group(1)) * 1_000_000) if m else 0
        if 'K' in s.upper():
            m = re.search(r'([\d.]+)[Kk]', s)
            return int(float(m.group(1)) * 1_000) if m else 0
        pure = re.sub(r'[^\d]', '', s)
        return int(pure) if pure else 0
    except Exception:
        return 0


def fmt_subscriber(n: int) -> str:
    """구독자 수를 보기 좋게 포맷합니다."""
    if n <= 0:
        return '-'
    if n >= 100_000_000:
        return f"{n/100_000_000:.1f}억"
    if n >= 10_000:
        return f"{n/10_000:.1f}만"
    if n >= 1_000:
        return f"{n/1_000:.1f}천"
    return str(n)


def scrape_subscriber_count(channel_url: str) -> int:
    """
    채널 페이지에서 구독자 수를 수집합니다.
    캐시를 먼저 확인하고, 없으면 크롤링합니다.
    """
    if not channel_url or not channel_url.startswith('http'):
        return 0

    cache = st.session_state.get('subscriber_cache', {})
    if channel_url in cache:
        return cache[channel_url]

    # base URL 정규화
    base_url = channel_url.rstrip('/')
    for suffix in ['/videos', '/shorts', '/streams', '/featured',
                   '/playlists', '/about', '/search']:
        if base_url.endswith(suffix):
            base_url = base_url[:-len(suffix)]
            break

    driver = get_driver()
    sub_count = 0
    try:
        driver.get(base_url)
        try:
            WebDriverWait(driver, 15).until(
                lambda d: 'ytInitialData' in d.page_source
            )
        except Exception:
            pass
        time.sleep(2)
        html = driver.page_source

        # ytInitialData에서 구독자 수 파싱
        data = _extract_yt_initial_data(html)
        if data:
            # header 내 subscriberCountText
            sub_raw = (
                _safe_get(data, 'header', 'pageHeaderRenderer',
                          'content', 'pageHeaderViewModel',
                          'metadata', 'contentMetadataViewModel',
                          'metadataRows', 1, 'metadataParts', 0, 'text', 'content') or
                _safe_get(data, 'header', 'c4TabbedHeaderRenderer',
                          'subscriberCountText', 'simpleText') or
                _safe_get(data, 'header', 'c4TabbedHeaderRenderer',
                          'subscriberCountText', 'runs', 0, 'text') or
                ''
            )
            if sub_raw:
                sub_count = parse_subscriber_count(sub_raw)

        # DOM fallback
        if sub_count == 0:
            patterns = [
                r'([\d.,]+[만천억MmKk]?\s*(?:만|천|억)?)\s*구독자',
                r'"subscriberCountText".*?"simpleText"\s*:\s*"([^"]+)"',
                r'구독자\s*([\d.,]+[만천억MmKk]*)',
            ]
            for pat in patterns:
                m = re.search(pat, html)
                if m:
                    sub_count = parse_subscriber_count(m.group(1))
                    if sub_count > 0:
                        break

    except Exception:
        pass
    finally:
        driver.quit()

    st.session_state.subscriber_cache[channel_url] = sub_count
    return sub_count


def scrape_subscribers_batch(channel_stats: pd.DataFrame,
                             progress_placeholder=None) -> pd.DataFrame:
    """
    channel_stats DataFrame의 각 채널에 대해 구독자 수를 수집하여
    'subscriber' 컬럼을 추가합니다.
    """
    total = len(channel_stats)
    subscribers = []

    for i, row in channel_stats.iterrows():
        ch_url = row.get('channel_url', '')
        if progress_placeholder:
            progress_placeholder.info(
                f"🔄 구독자 수 수집 중... ({i+1}/{total}) **{row['channel_name']}**"
            )
        sub = scrape_subscriber_count(ch_url)
        subscribers.append(sub)
        if i < total - 1:
            time.sleep(random.uniform(1.0, 2.0))

    channel_stats = channel_stats.copy()
    channel_stats['subscriber'] = subscribers
    return channel_stats


# ── 채널 통계 엑셀 생성 함수 ─────────────────────────
def build_channel_excel(channel_stats: pd.DataFrame, keyword: str) -> bytes:
    """
    채널 통계 DataFrame을 예쁘게 포맷된 엑셀 바이트로 반환합니다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "추천 채널"

    # ── 색상 정의 ────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # 진한 파랑
    ROW1_FILL = PatternFill("solid", fgColor="DEEAF1")   # 연한 파랑
    ROW2_FILL = PatternFill("solid", fgColor="FFFFFF")   # 흰색
    GOLD_FILL = PatternFill("solid", fgColor="FFD700")   # 금색 (1위)
    SILV_FILL = PatternFill("solid", fgColor="C0C0C0")   # 은색 (2위)
    BRNZ_FILL = PatternFill("solid", fgColor="CD7F32")   # 동색 (3위)

    HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    BODY_FONT = Font(name='Arial', size=10)
    BOLD_FONT = Font(name='Arial', bold=True, size=10)
    LINK_FONT = Font(name='Arial', size=10, color='0563C1', underline='single')

    center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right   = Alignment(horizontal='right',  vertical='center')

    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 제목 행 ──────────────────────────────────────
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"📺 '{keyword or '검색'}' 키워드 추천 채널 분석"
    title_cell.font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    title_cell.alignment = center
    title_cell.fill = PatternFill("solid", fgColor="EBF3FB")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:K2')
    ws['A2'].value = f"수집일: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}  |  총 {len(channel_stats)}개 채널"
    ws['A2'].font = Font(name='Arial', size=9, color='666666', italic=True)
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 18

    # ── 헤더 행 ──────────────────────────────────────
    headers = [
        ('순위', 6), ('채널명', 22), ('구독자', 12), ('영상 수', 9),
        ('평균 조회수', 12), ('최고 조회수', 12), ('총 조회수', 12),
        ('일반 영상', 9), ('쇼츠', 7), ('종합 점수', 10), ('채널 링크', 30),
    ]
    col_keys = [
        'rank', 'channel_name', 'subscriber', 'video_count',
        'avg_view', 'max_view', 'total_view',
        'regular_count', 'shorts_count', 'score', 'channel_url',
    ]

    for col_idx, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # ── 데이터 행 ─────────────────────────────────────
    def fmt_num(v):
        if v >= 100_000_000:
            return f"{v/100_000_000:.1f}억"
        if v >= 10_000:
            return f"{v/10_000:.1f}만"
        if v >= 1_000:
            return f"{v/1_000:.1f}천"
        return str(int(v)) if v else '0'

    medal_fills = {1: GOLD_FILL, 2: SILV_FILL, 3: BRNZ_FILL}

    for row_idx, (_, row) in enumerate(channel_stats.iterrows(), 4):
        rank = row.get('rank', row_idx - 3)
        fill = medal_fills.get(rank, ROW1_FILL if (row_idx % 2 == 0) else ROW2_FILL)
        ws.row_dimensions[row_idx].height = 20

        for col_idx, key in enumerate(col_keys, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.fill = fill

            val = row.get(key, '')

            if key == 'rank':
                cell.value = int(rank)
                cell.font = BOLD_FONT
                cell.alignment = center

            elif key == 'channel_name':
                cell.value = str(val)
                cell.font = BOLD_FONT if rank <= 3 else BODY_FONT
                cell.alignment = left

            elif key == 'subscriber':
                sub_n = int(val) if val else 0
                cell.value = fmt_subscriber(sub_n) if sub_n > 0 else '-'
                cell.font = BODY_FONT
                cell.alignment = right

            elif key in ('avg_view', 'max_view', 'total_view'):
                cell.value = fmt_num(int(val) if val else 0)
                cell.font = BODY_FONT
                cell.alignment = right

            elif key == 'score':
                cell.value = round(float(val), 1) if val else 0
                cell.font = BODY_FONT
                cell.alignment = center

            elif key == 'channel_url':
                url = str(val)
                if url.startswith('http'):
                    cell.value = url
                    cell.hyperlink = url
                    cell.font = LINK_FONT
                else:
                    cell.value = '-'
                    cell.font = BODY_FONT
                cell.alignment = left

            else:
                cell.value = int(val) if val else 0
                cell.font = BODY_FONT
                cell.alignment = center

    # ── 틀 고정 ──────────────────────────────────────
    ws.freeze_panes = 'A4'

    # ── 자동 필터 ─────────────────────────────────────
    ws.auto_filter.ref = f"A3:K{3 + len(channel_stats)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── 키워드 검색 스크래퍼 ───────────────────────────────
def scrape_keyword(keyword: str, upload_filter: str) -> pd.DataFrame:
    SEARCH_KEYWORD = keyword.replace(' ', '+')
    url_param = UPLOAD_FILTER_MAP.get(upload_filter, "")

    driver = get_driver()
    URL = f"https://www.youtube.com/results?search_query={SEARCH_KEYWORD}{url_param}"
    driver.get(URL)

    try:
        WebDriverWait(driver, 15).until(
            lambda d: 'ytInitialData' in d.page_source
        )
    except Exception:
        pass
    time.sleep(2)
    scroll(driver)

    html_source = driver.page_source
    driver.quit()

    data = _extract_yt_initial_data(html_source)
    if data:
        rows = _parse_all_renderers_from_data(data)
        if rows:
            return _rows_to_df(rows)

    soup = BeautifulSoup(html_source, 'html.parser')
    return _parse_search_results(soup)


# ── 채널 내 키워드 검색 ────────────────────────────────
def scrape_channel_search(channel_url: str, search_keyword: str,
                          channel_name: str = "") -> pd.DataFrame:
    base_url = channel_url.rstrip('/')
    for suffix in ['/videos', '/shorts', '/streams', '/featured', '/playlists', '/about', '/search']:
        if base_url.endswith(suffix):
            base_url = base_url[:-len(suffix)]
            break

    encoded_keyword = search_keyword.replace(' ', '+')
    search_url = f"{base_url}/search?query={encoded_keyword}"

    driver = get_driver()
    rows = []
    html_source = ""

    try:
        driver.get(search_url)

        try:
            WebDriverWait(driver, 20).until(
                lambda d: 'ytInitialData' in d.page_source
            )
        except Exception:
            pass
        time.sleep(3)

        try:
            btn = WebDriverWait(driver, 4).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(., 'Accept') or contains(., '동의') or contains(., '동의합니다')]")
                )
            )
            btn.click()
            time.sleep(1)
        except Exception:
            pass

        scroll(driver)
        html_source = driver.page_source

    finally:
        driver.quit()

    data = _extract_yt_initial_data(html_source)
    parse_method = 'none'
    if data:
        rows = _parse_channel_search_data(data)
        parse_method = 'itemSectionRenderer'
        if not rows:
            rows = _parse_all_renderers_from_data(data)
            parse_method = 'all_renderers_fallback'

    if not rows:
        soup = BeautifulSoup(html_source, 'html.parser')
        df_fallback = _parse_search_results(soup)
        parse_method = 'dom_fallback'
        if not df_fallback.empty:
            if channel_name:
                df_fallback['channel'] = channel_name
            _save_debug(channel_name or channel_url, search_url, len(df_fallback),
                        html_source, parse_method)
            return df_fallback

    _save_debug(channel_name or channel_url, search_url, len(rows),
                html_source, parse_method)

    df = _rows_to_df(rows)
    if channel_name:
        df['channel'] = channel_name

    return df


# ── 단일 채널 탭 스크래퍼 ────────────────────────────────
def scrape_channel(channel_url: str, tab: str = "동영상",
                   channel_name: str = "") -> pd.DataFrame:
    tab_map = {"동영상": "/videos", "쇼츠": "/shorts"}
    tab_suffix = tab_map.get(tab, "/videos")

    base_url = channel_url.rstrip('/')
    for suffix in ['/videos', '/shorts', '/streams', '/featured', '/playlists', '/about', '/search']:
        if base_url.endswith(suffix):
            base_url = base_url[:-len(suffix)]
            break
    target_url = base_url + tab_suffix

    driver = get_driver()
    try:
        driver.get(target_url)

        try:
            WebDriverWait(driver, 20).until(
                lambda d: 'ytInitialData' in d.page_source
            )
        except Exception:
            pass
        time.sleep(3)

        try:
            btn = WebDriverWait(driver, 4).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(., 'Accept') or contains(., '동의') or contains(., '동의합니다')]")
                )
            )
            btn.click()
            time.sleep(1)
        except Exception:
            pass

        scroll(driver)
        html_source = driver.page_source
    finally:
        driver.quit()

    rows = []
    data = _extract_yt_initial_data(html_source)

    if data:
        rows = _parse_all_renderers_from_data(data)

    if not rows:
        soup = BeautifulSoup(html_source, 'html.parser')
        df_fallback = _parse_channel_results(soup)
        if not df_fallback.empty:
            if channel_name:
                df_fallback['channel'] = channel_name
            return df_fallback

    df = _rows_to_df(rows)
    if channel_name:
        df['channel'] = channel_name

    _save_debug(channel_name or channel_url, target_url, len(rows),
                html_source, 'tab_collect')

    return df


# ── 다중 채널 수집 (탭 수집) ──────────────────────────────
def scrape_multiple_channels(channel_urls: list, channel_names: list,
                             tab: str = "동영상",
                             progress_cb=None) -> pd.DataFrame:
    all_dfs = []
    total = len(channel_urls)
    for i, (url, name) in enumerate(zip(channel_urls, channel_names)):
        if progress_cb:
            progress_cb(i, total, name)
        try:
            df = scrape_channel(url, tab=tab, channel_name=name)
            if not df.empty:
                all_dfs.append(df)
        except Exception as e:
            st.warning(f"⚠️ '{name}' 수집 실패: {e}")
        if i < total - 1:
            time.sleep(random.uniform(2, 4))

    if not all_dfs:
        df_empty = _empty_df()
        df_empty['channel'] = ''
        return df_empty

    result = pd.concat(all_dfs, ignore_index=True)
    result = result.drop_duplicates(subset=['link']).reset_index(drop=True)
    return result


# ── 다중 채널 내 키워드 검색 ─────────────────────────────
def scrape_multiple_channels_search(channel_urls: list, channel_names: list,
                                    search_keyword: str,
                                    progress_cb=None) -> pd.DataFrame:
    all_dfs = []
    total = len(channel_urls)
    for i, (url, name) in enumerate(zip(channel_urls, channel_names)):
        if progress_cb:
            progress_cb(i, total, name)
        try:
            df = scrape_channel_search(url, search_keyword, channel_name=name)
            if not df.empty:
                all_dfs.append(df)
                st.info(f"✅ '{name}': {len(df)}개 영상 수집")
            else:
                st.warning(f"⚠️ '{name}': 검색 결과 없음")
        except Exception as e:
            st.warning(f"⚠️ '{name}' 검색 실패: {e}")
        if i < total - 1:
            time.sleep(random.uniform(2, 4))

    if not all_dfs:
        df_empty = _empty_df()
        df_empty['channel'] = ''
        return df_empty

    result = pd.concat(all_dfs, ignore_index=True)
    result = result.drop_duplicates(subset=['link']).reset_index(drop=True)
    return result


# ── HTML 파싱: 검색 결과 ──────────────────────────────
def _parse_search_results(soup: BeautifulSoup) -> pd.DataFrame:
    renderers = soup.find_all('ytd-video-renderer')
    titles, links, views, dates = [], [], [], []
    channel_names_list, channel_urls_list = [], []
    for renderer in renderers:
        a_tag = renderer.find(class_='yt-simple-endpoint style-scope ytd-video-renderer')
        if not a_tag or not a_tag.get('href'):
            continue
        titles.append(a_tag.get_text().replace("\n", ""))
        links.append("https://youtube.com" + a_tag["href"])
        meta = renderer.find(class_='style-scope ytd-video-meta-block')
        raw = meta.get_text(separator='|').strip() if meta else ''
        views.append(_extract_view(raw))
        dates.append(_extract_date(raw))
        # 채널명 추출 시도
        ch_tag = renderer.find('yt-formatted-string', class_=lambda c: c and 'ytd-channel-name' in c)
        ch_name = ch_tag.get_text(strip=True) if ch_tag else ''
        ch_a = renderer.find('a', class_=lambda c: c and 'yt-simple-endpoint' in c and 'ytd-channel-name' in c)
        ch_url = ('https://youtube.com' + ch_a['href']) if ch_a and ch_a.get('href') else ''
        channel_names_list.append(ch_name)
        channel_urls_list.append(ch_url)
    return _build_df(titles, links, views, dates, channel_names_list, channel_urls_list)


def _parse_channel_results(soup: BeautifulSoup) -> pd.DataFrame:
    titles, links, views, dates = [], [], [], []
    channel_names_list, channel_urls_list = [], []

    items = soup.find_all('ytd-rich-item-renderer')
    if not items:
        items = soup.find_all('ytd-grid-video-renderer')

    for item in items:
        a_tag = item.find('a', id='video-title-link') or item.find('a', id='thumbnail')
        title_tag = item.find('yt-formatted-string', id='video-title') or \
                    item.find('span', id='video-title')

        title = title_tag.get_text(strip=True) if title_tag else ""
        link  = ("https://youtube.com" + a_tag["href"]) if (a_tag and a_tag.get("href")) else ""

        if not link:
            continue

        meta_block = item.find(class_='style-scope ytd-video-meta-block')
        raw = meta_block.get_text(separator='|').strip() if meta_block else ""

        view_spans = item.find_all('span', class_=lambda c: c and 'ytd-grid-video-renderer' in c)
        if not raw:
            raw = '|'.join(s.get_text(strip=True) for s in view_spans)

        titles.append(title)
        links.append(link)
        views.append(_extract_view(raw))
        dates.append(_extract_date(raw))
        channel_names_list.append('')
        channel_urls_list.append('')

    if not titles:
        return _parse_search_results(soup)

    return _build_df(titles, links, views, dates, channel_names_list, channel_urls_list)


# ── 공통 파싱 헬퍼 ────────────────────────────────────
def _extract_view(text: str) -> str:
    if not text:
        return ''
    text = re.sub(r'[•\n\r]', '|', text)

    m = re.search(r'조회수\s*([\d.,]+\s*(?:[만천억])?)', text)
    if m:
        return re.sub(r'[회\s]', '', m.group(1)).strip()

    m = re.search(r'([\d.]+\s*[만천억])', text)
    if m:
        return re.sub(r'[\s]', '', m.group(1)).strip()

    m = re.search(r'(\d{1,3}(?:,\d{3})+)', text)
    if m:
        return m.group(1).strip()

    m = re.search(r'(\d{5,})', text)
    if m:
        return m.group(1).strip()

    return ''


def _extract_date(text: str) -> str:
    m = re.search(r'(\d+\s*(?:분|시간|일|주|개월|년)\s*전)', text)
    if m:
        return m.group(1).strip()
    return ''


def _empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=[
        'title', 'link', 'view', 'upload_date',
        'view_num', 'days_ago', 'is_shorts',
        'channel_name', 'channel_url'
    ])


def _build_df(titles, links, views, dates,
              channel_names=None, channel_urls=None) -> pd.DataFrame:
    min_len = min(len(titles), len(links), len(views), len(dates))
    if min_len == 0:
        return _empty_df()
    channel_names = channel_names or [''] * min_len
    channel_urls  = channel_urls  or [''] * min_len
    df = pd.DataFrame({
        'title':        [str(t) for t in titles[:min_len]],
        'link':         [str(l) for l in links[:min_len]],
        'view':         [str(v) for v in views[:min_len]],
        'upload_date':  [str(d) for d in dates[:min_len]],
        'channel_name': [str(c) for c in channel_names[:min_len]],
        'channel_url':  [str(u) for u in channel_urls[:min_len]],
    })
    df['view_num'] = df['view'].apply(parse_view_count)
    df['days_ago'] = df['upload_date'].apply(date_to_days)
    df['is_shorts'] = [
        is_shorts(str(row['link']), str(row['title']))
        for _, row in df.iterrows()
    ]
    return df


def _rows_to_df(rows: list) -> pd.DataFrame:
    if not rows:
        return _empty_df()

    clean_rows = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        clean_rows.append({
            'title':        str(r.get('title', '')),
            'link':         str(r.get('link', '')),
            'view':         str(r.get('view', '')),
            'upload_date':  str(r.get('upload_date', '')),
            'channel_name': str(r.get('channel_name', '')),
            'channel_url':  str(r.get('channel_url', '')),
        })

    if not clean_rows:
        return _empty_df()

    df = pd.DataFrame(clean_rows)
    df['view_num'] = df['view'].apply(parse_view_count)
    df['days_ago'] = df['upload_date'].apply(date_to_days)
    df['is_shorts'] = [
        is_shorts(str(row['link']), str(row['title']))
        for _, row in df.iterrows()
    ]
    return df


# ── 콘텐츠 유형 필터 ─────────────────────────────────
def filter_content_type(df: pd.DataFrame, content_type: str) -> pd.DataFrame:
    if content_type == "일반 영상만":
        return df[~df['is_shorts']].copy()
    elif content_type == "쇼츠만":
        return df[df['is_shorts']].copy()
    return df.copy()


# ── AI 키워드 추천 ────────────────────────────────────
def get_ai_keyword_recommendations(df: pd.DataFrame, search_keyword: str) -> str:
    top_df = df.nlargest(50, 'view_num')[['title', 'view', 'view_num']]
    titles_data = top_df.to_dict(orient='records')

    prompt = f"""당신은 유튜브 콘텐츠 전략 전문가입니다.
아래는 유튜브에서 '{search_keyword}' 키워드로 검색했을 때 조회수가 높은 영상 제목 목록입니다.

{json.dumps(titles_data, ensure_ascii=False, indent=2)}

위 데이터를 분석하여 다음을 JSON 형식으로 답변해 주세요. JSON만 출력하고 다른 텍스트는 절대 포함하지 마세요.

{{
  "top_keywords": [
    {{
      "keyword": "키워드",
      "reason": "이 키워드가 조회수에 효과적인 이유 (1~2문장)",
      "example_titles": ["실제로 잘 터진 제목 예시 1", "예시 2"],
      "avg_view": "해당 키워드 포함 영상 평균 조회수 (예: 약 5만)"
    }}
  ],
  "recommended_title_patterns": [
    {{
      "pattern": "제목 패턴 (예: [충격] {{주제}} 했더니 {{결과}})",
      "reason": "이 패턴이 효과적인 이유"
    }}
  ],
  "insight": "전체 데이터에서 발견한 핵심 인사이트 (3~4문장)"
}}

top_keywords는 5개, recommended_title_patterns는 3개를 추천해 주세요."""

    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}]
    )
    return response.content[0].text


# ── ★ NEW: 채널 통계 집계 함수 ────────────────────────
def extract_channel_stats(df: pd.DataFrame) -> pd.DataFrame:
    """
    수집된 영상 데이터에서 채널별 통계를 집계합니다.
    channel_name 컬럼을 기준으로 그룹핑합니다.
    """
    if 'channel_name' not in df.columns:
        return pd.DataFrame()

    # 채널명이 있는 행만 사용
    ch_df = df[df['channel_name'].str.strip() != ''].copy()
    if ch_df.empty:
        return pd.DataFrame()

    agg = ch_df.groupby('channel_name').agg(
        video_count=('link', 'count'),
        avg_view=('view_num', 'mean'),
        max_view=('view_num', 'max'),
        total_view=('view_num', 'sum'),
        channel_url=('channel_url', 'first'),
        shorts_count=('is_shorts', 'sum'),
    ).reset_index()

    agg['avg_view'] = agg['avg_view'].astype(int)
    agg['max_view'] = agg['max_view'].astype(int)
    agg['total_view'] = agg['total_view'].astype(int)
    agg['shorts_count'] = agg['shorts_count'].astype(int)
    agg['regular_count'] = agg['video_count'] - agg['shorts_count']

    # 구독자 수는 별도 수집 전까지 0으로 초기화
    if 'subscriber' not in agg.columns:
        agg['subscriber'] = 0

    # 점수: 평균 조회수 60% + 영상 수 20% + 최고 조회수 20%
    max_avg = agg['avg_view'].max() or 1
    max_cnt = agg['video_count'].max() or 1
    max_max = agg['max_view'].max() or 1
    agg['score'] = (
        (agg['avg_view'] / max_avg) * 0.6 +
        (agg['video_count'] / max_cnt) * 0.2 +
        (agg['max_view'] / max_max) * 0.2
    ) * 100

    agg = agg.sort_values('score', ascending=False).reset_index(drop=True)
    agg['rank'] = agg.index + 1

    return agg


# ── ★ NEW: AI 추천 채널 분석 함수 ─────────────────────
def get_ai_channel_recommendations(channel_stats: pd.DataFrame,
                                   search_keyword: str) -> str:
    """
    채널별 통계를 바탕으로 AI가 추천 채널을 분석합니다.
    """
    top_channels = channel_stats.head(20).to_dict(orient='records')

    prompt = f"""당신은 유튜브 채널 분석 전문가입니다.
아래는 '{search_keyword}' 키워드 검색 결과에서 수집된 채널별 통계입니다.

{json.dumps(top_channels, ensure_ascii=False, indent=2)}

각 채널의 video_count(영상 수), avg_view(평균 조회수), max_view(최고 조회수), total_view(총 조회수), score(종합 점수)를 종합 분석하여 다음 JSON을 반환하세요. JSON만 출력하고 다른 텍스트는 절대 포함하지 마세요.

{{
  "recommended_channels": [
    {{
      "channel_name": "채널명",
      "reason": "이 채널을 추천하는 구체적 이유 (데이터 근거 포함, 2~3문장)",
      "strength": "채널의 핵심 강점 (예: 꾸준한 업로드, 높은 평균 조회수 등)",
      "content_strategy": "이 채널에서 배울 수 있는 콘텐츠 전략 팁",
      "avg_view_label": "평균 조회수 표시 (예: 약 12만)"
    }}
  ],
  "market_insight": "이 키워드 분야의 채널 생태계에 대한 핵심 인사이트 (3~4문장)",
  "entry_strategy": "신규 채널이 이 키워드로 성장하기 위한 전략 조언 (3~4문장)"
}}

recommended_channels는 최대 5개를 선정해 주세요. 단순히 조회수가 높은 채널보다, 실질적으로 벤치마킹할 가치가 있는 채널을 우선 추천해 주세요."""

    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1500,
        messages=[{"role": "user", "content": prompt}]
    )
    return response.content[0].text


# ── 결과 표시 함수 ────────────────────────────────────
def render_results():
    df = st.session_state.df
    keyword = st.session_state.search_keyword
    total_collected = st.session_state.total_collected
    filtered_count = st.session_state.filtered_count

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📦 전체 수집", f"{total_collected}개")
    col2.metric("✅ 필터 후",  f"{filtered_count}개")
    col3.metric("🩳 쇼츠",    f"{df['is_shorts'].sum()}개")
    ch_count = df['channel'].nunique() if 'channel' in df.columns else "-"
    col4.metric("📺 채널 수",  f"{ch_count}개" if 'channel' in df.columns else (keyword if keyword else "전체"))

    if filtered_count == 0:
        st.warning("⚠️ 필터 조건에 맞는 영상이 없습니다. 조건을 완화해 보세요.")
        return

    st.markdown("---")
    st.markdown("#### 🔧 필터 및 정렬")

    filter_col1, filter_col2, filter_col3 = st.columns([2, 2, 2])

    with filter_col1:
        sort_option = st.radio(
            "🔃 정렬 기준",
            options=["기본 순서 (관련도)", "조회수 많은 순", "등록일 최신순"],
            horizontal=False,
            key="sort_radio"
        )

    with filter_col2:
        in_tab_type = st.radio(
            "🩳 콘텐츠 유형",
            options=["전체", "일반 영상만", "쇼츠만"],
            horizontal=False,
            key="in_tab_type"
        )

    view_min_all = int(df['view_num'].min())
    view_max_all = int(df['view_num'].max())

    with filter_col3:
        if view_min_all < view_max_all:
            st.markdown("**👁️ 조회수 범위**")
            view_range = st.slider(
                "조회수 범위",
                min_value=view_min_all,
                max_value=view_max_all,
                value=(view_min_all, view_max_all),
                step=max(1, (view_max_all - view_min_all) // 100),
                format="%d",
                key="view_range_slider",
                label_visibility="collapsed"
            )
            view_lo, view_hi = view_range
        else:
            view_lo, view_hi = view_min_all, view_max_all

    df_sorted = df.copy()
    df_sorted = df_sorted[
        (df_sorted['view_num'] >= view_lo) &
        (df_sorted['view_num'] <= view_hi)
    ]
    df_sorted = filter_content_type(df_sorted, in_tab_type)

    if sort_option == "조회수 많은 순":
        df_sorted = df_sorted.sort_values('view_num', ascending=False)
    elif sort_option == "등록일 최신순":
        df_sorted = df_sorted.sort_values('days_ago', ascending=True)

    st.markdown(f"<span style='color:gray;font-size:0.9em'>총 {len(df_sorted)}개 영상</span>",
                unsafe_allow_html=True)
    st.markdown("---")

    # ── ★ 탭에 "추천 채널" 추가 ───────────────────────
    tab1, tab2, tab3 = st.tabs(["📊 수집 결과", "📺 추천 채널", "🤖 AI 키워드 추천"])

    with tab1:
        if len(df_sorted) == 0:
            st.warning("⚠️ 해당 조건에 맞는 영상이 없습니다. 조건을 변경해 보세요.")
        else:
            drop_cols = ['view_num', 'days_ago', '_debug', 'channel_name', 'channel_url']
            df_display = df_sorted.drop(columns=[c for c in drop_cols if c in df_sorted.columns]).copy()
            df_display['is_shorts'] = df_display['is_shorts'].apply(
                lambda x: "🩳 쇼츠" if x else "🎬 일반"
            )
            df_display = df_display.rename(columns={'is_shorts': '유형'})
            if 'channel' in df_display.columns:
                cols = ['channel', 'title', '유형', 'view', 'upload_date', 'link']
                df_display = df_display[[c for c in cols if c in df_display.columns]]
                df_display = df_display.rename(columns={'channel': '채널'})
            df_display['link'] = df_display['link'].apply(
                lambda x: f'<a href="{x}" target="_blank">🔗 보기</a>'
            )
            st.write(df_display.to_html(escape=False, index=False), unsafe_allow_html=True)

        st.divider()
        drop_cols_csv = ['view_num', 'days_ago', '_debug']
        csv_df = df_sorted.drop(columns=[c for c in drop_cols_csv if c in df_sorted.columns]).copy()
        csv_df['is_shorts'] = csv_df['is_shorts'].apply(lambda x: "쇼츠" if x else "일반")
        csv = csv_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
        st.download_button(
            label="⬇️ CSV 다운로드",
            data=csv,
            file_name=f"youtube_{keyword or 'channel'}.csv",
            mime="text/csv",
            type="primary"
        )

    # ── ★ NEW: 추천 채널 탭 ───────────────────────────
    with tab2:
        st.subheader("📺 키워드 추천 채널")
        st.caption(f"수집된 영상 데이터를 기반으로 **'{keyword or '검색'}'** 키워드에서 주목할 만한 채널을 분석합니다.")

        # 채널 통계 집계
        channel_stats = extract_channel_stats(df_sorted)

        if channel_stats.empty:
            st.info(
                "ℹ️ 채널 정보가 수집되지 않았습니다.\n\n"
                "채널 추천 기능은 **키워드 검색** 모드에서 가장 잘 동작합니다. "
                "키워드로 검색하면 각 영상의 채널 정보가 함께 수집됩니다."
            )
        else:
            # ── 구독자 수 수집 버튼 ───────────────────
            st.markdown("### 👥 구독자 수 수집")

            # 캐시에서 구독자 수 복원
            cached_subs = st.session_state.get('subscriber_cache', {})
            if cached_subs:
                channel_stats['subscriber'] = channel_stats['channel_url'].apply(
                    lambda u: cached_subs.get(u, 0)
                )
                has_subscribers = channel_stats['subscriber'].sum() > 0
            else:
                has_subscribers = False

            sub_col1, sub_col2 = st.columns([2, 4])
            with sub_col1:
                fetch_sub_btn = st.button(
                    "🔍 구독자 수 수집 시작",
                    type="primary",
                    use_container_width=True,
                    key="fetch_sub_btn",
                    help="각 채널 페이지를 방문하여 구독자 수를 수집합니다. 채널 수에 따라 시간이 소요됩니다."
                )
            with sub_col2:
                if has_subscribers:
                    st.success(f"✅ {channel_stats[channel_stats['subscriber'] > 0]['channel_name'].count()}개 채널 구독자 수 수집 완료")
                else:
                    st.caption("⚠️ 구독자 수를 수집하면 구독자순 정렬이 가능합니다. 채널 수에 따라 1~3분 소요됩니다.")

            if fetch_sub_btn:
                sub_progress = st.empty()
                with st.spinner("채널 페이지에서 구독자 수를 수집하고 있습니다..."):
                    channel_stats = scrape_subscribers_batch(channel_stats, sub_progress)
                    # 캐시 업데이트
                    for _, row in channel_stats.iterrows():
                        url = row.get('channel_url', '')
                        if url:
                            st.session_state.subscriber_cache[url] = row.get('subscriber', 0)
                sub_progress.empty()
                has_subscribers = channel_stats['subscriber'].sum() > 0
                if has_subscribers:
                    st.success("✅ 구독자 수 수집 완료!")
                else:
                    st.warning("⚠️ 구독자 수를 가져오지 못했습니다. 채널 URL을 확인해 주세요.")

            st.divider()

            # ── 정렬 기준 선택 ────────────────────────
            st.markdown("### 📊 채널별 영상 현황")

            sort_col, _ = st.columns([2, 4])
            with sort_col:
                sort_options = ["종합 점수순", "평균 조회수순", "최고 조회수순", "총 조회수순", "영상 수순"]
                if has_subscribers:
                    sort_options.insert(1, "구독자 수순")
                ch_sort = st.selectbox(
                    "🔃 정렬 기준",
                    options=sort_options,
                    key="ch_sort_select"
                )

            sort_map = {
                "종합 점수순":    'score',
                "구독자 수순":   'subscriber',
                "평균 조회수순": 'avg_view',
                "최고 조회수순": 'max_view',
                "총 조회수순":   'total_view',
                "영상 수순":     'video_count',
            }
            sort_col_key = sort_map.get(ch_sort, 'score')
            channel_stats_sorted = channel_stats.sort_values(
                sort_col_key, ascending=False
            ).reset_index(drop=True)
            channel_stats_sorted['rank'] = channel_stats_sorted.index + 1

            # ── 채널 통계 표 ──────────────────────────
            display_cols = ['rank', 'channel_name']
            display_rename = {'rank': '순위', 'channel_name': '채널명'}

            if has_subscribers:
                display_cols.append('subscriber')
                display_rename['subscriber'] = '구독자'

            display_cols += ['video_count', 'avg_view', 'max_view',
                             'total_view', 'regular_count', 'shorts_count',
                             'score', 'channel_url']
            display_rename.update({
                'video_count':   '영상 수',
                'avg_view':      '평균 조회수',
                'max_view':      '최고 조회수',
                'total_view':    '총 조회수',
                'regular_count': '일반',
                'shorts_count':  '쇼츠',
                'score':         '종합 점수',
                'channel_url':   '링크',
            })

            display_stats = channel_stats_sorted[
                [c for c in display_cols if c in channel_stats_sorted.columns]
            ].copy()

            def fmt_view(v):
                v = int(v) if v else 0
                if v >= 100_000_000:
                    return f"{v/100_000_000:.1f}억"
                elif v >= 10_000:
                    return f"{v/10_000:.1f}만"
                elif v >= 1_000:
                    return f"{v/1_000:.1f}천"
                return str(v)

            for col in ['avg_view', 'max_view', 'total_view']:
                if col in display_stats.columns:
                    display_stats[col] = display_stats[col].apply(fmt_view)

            if 'subscriber' in display_stats.columns:
                display_stats['subscriber'] = display_stats['subscriber'].apply(
                    lambda v: fmt_subscriber(int(v)) if v else '-'
                )

            if 'score' in display_stats.columns:
                display_stats['score'] = display_stats['score'].apply(
                    lambda v: f"{v:.1f}" if v else '0'
                )

            display_stats['channel_url'] = display_stats['channel_url'].apply(
                lambda x: f'<a href="{x}" target="_blank">🔗 채널</a>'
                if x and str(x).startswith('http') else ''
            )
            display_stats = display_stats.rename(columns=display_rename)
            st.write(display_stats.to_html(escape=False, index=False), unsafe_allow_html=True)

            # ── 엑셀 다운로드 ─────────────────────────
            st.divider()
            st.markdown("#### ⬇️ 채널 분석 데이터 다운로드")
            dl_col1, dl_col2 = st.columns([1, 3])
            with dl_col1:
                try:
                    excel_bytes = build_channel_excel(channel_stats_sorted, keyword)
                    st.download_button(
                        label="📥 엑셀(.xlsx) 다운로드",
                        data=excel_bytes,
                        file_name=f"추천채널_{keyword or 'channel'}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"엑셀 생성 오류: {e}")
            with dl_col2:
                st.caption(
                    "채널명, 구독자 수, 영상 수, 평균/최고/총 조회수, 종합 점수, 채널 링크가 포함됩니다.\n"
                    "구독자 수를 먼저 수집하면 더 완성된 데이터를 받을 수 있습니다."
                )

            st.divider()

            # ── AI 추천 채널 분석 ─────────────────────
            st.markdown("### 🤖 AI 채널 추천 분석")

            # 캐시 확인 (같은 키워드면 재사용)
            cache_hit = (
                st.session_state.channel_recommendations is not None and
                st.session_state.channel_rec_keyword == (keyword or '')
            )

            if cache_hit:
                result_json = st.session_state.channel_recommendations
                _render_channel_rec(result_json, keyword)
            else:
                col_btn, col_info = st.columns([1, 3])
                with col_btn:
                    analyze_btn = st.button(
                        "🔍 AI 채널 분석 시작",
                        type="primary",
                        use_container_width=True,
                        key="analyze_channel_btn"
                    )
                with col_info:
                    st.caption("AI가 채널 통계를 분석해 벤치마킹할 채널과 콘텐츠 전략을 추천합니다.")

                if analyze_btn:
                    with st.spinner("🧠 AI가 채널 데이터를 분석 중입니다..."):
                        try:
                            raw = get_ai_channel_recommendations(channel_stats_sorted, keyword or "검색")
                            raw_clean = re.sub(r'```json|```', '', raw).strip()
                            result_json = json.loads(raw_clean)
                            st.session_state.channel_recommendations = result_json
                            st.session_state.channel_rec_keyword = keyword or ''
                            _render_channel_rec(result_json, keyword)
                        except json.JSONDecodeError:
                            st.markdown(raw)
                        except Exception as e:
                            st.error(f"❌ AI 분석 오류: {e}")

    with tab3:
        st.subheader("🤖 AI 키워드 & 제목 패턴 추천")
        st.caption("조회수 상위 영상들을 AI가 분석하여 잘 터지는 키워드와 제목 패턴을 추천합니다.")

        with st.spinner("🧠 AI가 데이터를 분석 중입니다..."):
            try:
                raw = get_ai_keyword_recommendations(df, keyword or "채널 영상")
                raw_clean = re.sub(r'```json|```', '', raw).strip()
                result = json.loads(raw_clean)

                st.info(f"💡 **핵심 인사이트**\n\n{result.get('insight', '')}")

                st.markdown("### 🔑 잘 터지는 키워드 TOP 5")
                for i, kw in enumerate(result.get('top_keywords', []), 1):
                    with st.expander(f"#{i}  **{kw['keyword']}**  —  평균 조회수 {kw.get('avg_view', '-')}"):
                        st.markdown(f"**이유:** {kw['reason']}")
                        st.markdown("**실제 잘 터진 제목 예시:**")
                        for ex in kw.get('example_titles', []):
                            st.markdown(f"- {ex}")

                st.markdown("### ✍️ 추천 제목 패턴")
                for p in result.get('recommended_title_patterns', []):
                    st.markdown(f"**패턴:** `{p['pattern']}`")
                    st.markdown(f"→ {p['reason']}")
                    st.divider()

            except json.JSONDecodeError:
                st.markdown(raw)
            except Exception as e:
                st.error(f"❌ AI 분석 오류: {e}")


def _render_channel_rec(result_json: dict, keyword: str):
    """AI 채널 추천 결과를 렌더링합니다."""
    st.success("✅ AI 분석 완료")

    # 시장 인사이트
    market = result_json.get('market_insight', '')
    if market:
        st.info(f"🌏 **시장 인사이트**\n\n{market}")

    # 추천 채널 카드
    st.markdown("#### 🏆 벤치마킹 추천 채널")
    rec_channels = result_json.get('recommended_channels', [])

    for i, ch in enumerate(rec_channels, 1):
        medal = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"][i - 1] if i <= 5 else f"#{i}"
        ch_name = ch.get('channel_name', '알 수 없음')
        avg_label = ch.get('avg_view_label', '')

        with st.expander(
            f"{medal} **{ch_name}**" + (f"  —  평균 {avg_label}" if avg_label else ""),
            expanded=(i == 1)
        ):
            col_l, col_r = st.columns([1, 1])
            with col_l:
                st.markdown(f"**💬 추천 이유**\n\n{ch.get('reason', '')}")
                st.markdown(f"**⭐ 핵심 강점**\n\n{ch.get('strength', '')}")
            with col_r:
                st.markdown(f"**📌 콘텐츠 전략 팁**\n\n{ch.get('content_strategy', '')}")

    # 진입 전략
    entry = result_json.get('entry_strategy', '')
    if entry:
        st.markdown("#### 🚀 신규 채널 진입 전략")
        st.warning(entry)

    # 재분석 버튼
    if st.button("🔄 재분석", key="re_analyze_ch_btn"):
        st.session_state.channel_recommendations = None
        st.session_state.channel_rec_keyword = ''
        st.rerun()


# ── 실행 ─────────────────────────────────────────────
if run_btn:
    # 크롤링 실행 시 채널 추천 캐시 초기화
    st.session_state.channel_recommendations = None
    st.session_state.channel_rec_keyword = ''

    if search_mode == "키워드 검색" and not keyword:
        st.warning("⚠️ 검색 키워드를 입력해주세요!")

    elif search_mode == "채널 검색" and not st.session_state.selected_channels:
        st.warning("⚠️ 수집할 채널을 1개 이상 선택해주세요!")

    elif (search_mode == "채널 검색" and
          st.session_state.get("channel_search_mode") == "키워드로 채널 내 검색" and
          not keyword):
        st.warning("⚠️ 채널 내 검색할 키워드를 입력해주세요!")

    else:
        st.session_state.scrape_debug = {}

        try:
            if search_mode == "키워드 검색":
                with st.spinner(f"🔄 '{keyword}' 크롤링 중..."):
                    df = scrape_keyword(keyword, upload_filter)

            else:
                sel_urls  = st.session_state.selected_channels
                name_map  = {c['url']: c['name'] for c in st.session_state.channel_list}
                sel_names = [name_map.get(u, u) for u in sel_urls]

                prog_bar  = st.progress(0, text="채널 수집 준비 중...")
                status_tx = st.empty()

                def progress_cb(i, total, name):
                    pct = int(i / total * 100)
                    prog_bar.progress(pct, text=f"({i+1}/{total}) '{name}' 수집 중...")
                    status_tx.info(f"🔄 **{name}** 채널 크롤링 중...")

                ch_mode = st.session_state.get("channel_search_mode", "탭 수집 (전체 영상)")

                if ch_mode == "키워드로 채널 내 검색":
                    st.info(f"🔍 각 채널에서 **'{keyword}'** 키워드로 검색합니다.")
                    df = scrape_multiple_channels_search(
                        sel_urls, sel_names,
                        search_keyword=keyword,
                        progress_cb=progress_cb
                    )
                else:
                    tab_sel = channel_tab if content_type != "쇼츠만" else "쇼츠"
                    df = scrape_multiple_channels(
                        sel_urls, sel_names,
                        tab=tab_sel,
                        progress_cb=progress_cb
                    )

                prog_bar.progress(100, text="✅ 수집 완료!")
                status_tx.empty()

            total_collected = len(df)

            if search_mode == "키워드 검색":
                day_limit = UPLOAD_DAY_LIMIT.get(upload_filter, 9999)
                if day_limit < 9999:
                    df = df[df['days_ago'] <= day_limit]

            ch_mode = st.session_state.get("channel_search_mode", "탭 수집 (전체 영상)")

            if search_mode == "키워드 검색" or ch_mode == "탭 수집 (전체 영상)":
                df = filter_content_type(df, content_type)

            if (search_mode == "채널 검색" and
                ch_mode == "탭 수집 (전체 영상)" and
                keyword):
                df = df[df['title'].str.contains(keyword, case=False, na=False)]

            if use_view_filter and min_view > 0:
                df = df[df['view_num'] >= min_view]

            if df.empty or 'title' not in df.columns:
                st.warning("⚠️ 영상 데이터를 가져오지 못했습니다. 채널명/키워드를 확인하거나 잠시 후 다시 시도해주세요.")
            else:
                st.session_state.df = df.reset_index(drop=True)
                st.session_state.search_keyword = keyword or ""
                st.session_state.total_collected = total_collected
                st.session_state.filtered_count = len(df)

            if st.session_state.get('scrape_debug'):
                with st.expander("🔍 디버그 정보 (개발자용)", expanded=df.empty):
                    st.json(st.session_state.scrape_debug)

        except Exception as e:
            st.error(f"❌ 오류 발생: {e}")
            st.info("크롬 드라이버나 패키지 문제일 수 있습니다. 터미널 오류를 확인해주세요.")

# ── session_state에 결과가 있으면 항상 표시 ──────────────
if st.session_state.df is not None:
    render_results()
